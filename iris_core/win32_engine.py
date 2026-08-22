"""
IRIS Core: Low-Level Win32 & COM Execution Engine
Provides native Windows OS control, message queuing, window lifecycle management,
and Microsoft Excel COM integration for fast, zero-mouse cross-app automation.
"""

import sys
import os
import time
import subprocess
import ctypes
from ctypes import wintypes
from typing import List, Dict, Any, Optional

# Setup Win32 User32 CTypes API Signatures for 64-bit robustness
user32 = ctypes.windll.user32
user32.OpenDesktopW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD, wintypes.BOOL, wintypes.DWORD]
user32.OpenDesktopW.restype = wintypes.HDESK
user32.SetThreadDesktop.argtypes = [wintypes.HDESK]
user32.SetThreadDesktop.restype = wintypes.BOOL
user32.FindWindowExW.argtypes = [wintypes.HWND, wintypes.HWND, wintypes.LPCWSTR, wintypes.LPCWSTR]
user32.FindWindowExW.restype = wintypes.HWND
user32.GetWindowTextW.argtypes = [wintypes.HWND, wintypes.LPWSTR, ctypes.c_int]
user32.GetWindowTextW.restype = ctypes.c_int
user32.GetWindowTextLengthW.argtypes = [wintypes.HWND]
user32.GetWindowTextLengthW.restype = ctypes.c_int
user32.GetClassNameW.argtypes = [wintypes.HWND, wintypes.LPWSTR, ctypes.c_int]
user32.GetClassNameW.restype = ctypes.c_int
user32.IsWindowVisible.argtypes = [wintypes.HWND]
user32.IsWindowVisible.restype = wintypes.BOOL
user32.IsIconic.argtypes = [wintypes.HWND]
user32.IsIconic.restype = wintypes.BOOL
user32.IsZoomed.argtypes = [wintypes.HWND]
user32.IsZoomed.restype = wintypes.BOOL

def ensure_interactive_desktop():
    """
    Ensures the calling thread is attached to the interactive Windows desktop (WinSta0\\Default)
    and has COM initialized, allowing background worker threads to access UIA and HWNDs.
    """
    try:
        hdesk = user32.OpenDesktopW('Default', 0, False, 0x10000000)
        if hdesk:
            user32.SetThreadDesktop(hdesk)
    except Exception:
        pass
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except Exception:
        pass

# Ensure desktop attached immediately upon module load before GUI hooks initialize
ensure_interactive_desktop()

import win32gui
import win32con
import win32api
import win32process
import psutil
import pyperclip
import pyautogui
pyautogui.FAILSAFE = False

def get_all_desktop_windows() -> List[Dict[str, Any]]:
    """
    Safely and reliably enumerates all interactive top-level desktop application windows.
    Bypasses pywin32 EnumWindows buffer limitations by using Win32 FindWindowEx.
    """
    ensure_interactive_desktop()

    cur = 0
    windows = []
    while True:
        cur = user32.FindWindowExW(0, cur, None, None)
        if not cur:
            break
        if not user32.IsWindowVisible(cur):
            continue
        length = user32.GetWindowTextLengthW(cur)
        if length == 0:
            continue
        buff = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(cur, buff, length + 1)
        title = buff.value.strip()
        if not title:
            continue

        cls_buff = ctypes.create_unicode_buffer(256)
        user32.GetClassNameW(cur, cls_buff, 256)
        cls = cls_buff.value

        # Filter out OS desktop infrastructure and invisible background hosts
        if cls in ['Progman', 'Shell_TrayWnd', 'NotifyIconOverflowWindow', 'Windows.UI.Core.CoreWindow', 'WorkerW']:
            continue

        rect = win32gui.GetWindowRect(cur)
        w = rect[2] - rect[0]
        h = rect[3] - rect[1]
        if w < 100 or h < 100:
            continue

        pid = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(cur, ctypes.byref(pid))
        pname = ''
        exe_path = ''
        if pid.value:
            try:
                proc = psutil.Process(pid.value)
                pname = proc.name().lower()
                exe_path = proc.exe()
            except Exception:
                pass

        # Exclude internal IRIS Electron UI windows (e.g. transparent floating overlay)
        if pname in ['electron.exe', 'iris.exe'] and (title.lower() in ['hackathon-iris', 'iris'] or not title):
            continue

        is_min = bool(user32.IsIconic(cur))
        is_max = bool(user32.IsZoomed(cur))

        windows.append({
            'hwnd': cur,
            'title': title,
            'class': cls,
            'pid': pid.value,
            'pname': pname,
            'exe': exe_path,
            'rect': (rect[0], rect[1], w, h),
            'is_min': is_min,
            'is_max': is_max
        })
    return windows

KNOWN_APP_PROCESSES = {
    "vscode": ["code.exe", "antigravity.exe", "antigravity ide.exe", "vscodium.exe"],
    "code": ["code.exe", "antigravity.exe", "antigravity ide.exe", "vscodium.exe"],
    "visual studio code": ["code.exe", "antigravity.exe", "antigravity ide.exe", "vscodium.exe"],
    "chrome": ["chrome.exe", "msedge.exe", "brave.exe", "firefox.exe"],
    "google chrome": ["chrome.exe", "msedge.exe", "brave.exe", "firefox.exe"],
    "browser": ["chrome.exe", "msedge.exe", "brave.exe", "firefox.exe"],
    "terminal": ["windowsterminal.exe", "wt.exe", "powershell.exe", "pwsh.exe", "cmd.exe"],
    "wt": ["windowsterminal.exe", "wt.exe"],
    "powershell": ["powershell.exe", "pwsh.exe"],
    "cmd": ["cmd.exe"],
    "notepad": ["notepad.exe", "notepad++.exe"],
    "excel": ["excel.exe"],
    "spotify": ["spotify.exe"],
    "discord": ["discord.exe"],
    "slack": ["slack.exe"],
    "explorer": ["explorer.exe"]
}

def find_window_by_name(keyword: str, must_be_visible: bool = False) -> Optional[int]:
    """
    Finds a top-level window whose process, class, or title matches the keyword.
    Uses exact process matching and smart aliases to prevent opening extra terminals.
    """
    if not keyword:
        return None

    windows = get_all_desktop_windows()
    kw = keyword.lower().strip().replace(".exe", "")

    # Pass 1: Known process name mapping
    target_procs = KNOWN_APP_PROCESSES.get(kw)
    if target_procs:
        for win in windows:
            if win['pname'] in target_procs:
                if kw == "explorer" and win['class'] != "CabinetWClass":
                    continue
                return win['hwnd']

    # Pass 2: Process name exact match
    for win in windows:
        p_base = win['pname'].replace(".exe", "")
        if p_base == kw or f"{kw}.exe" == win['pname']:
            return win['hwnd']

    # Pass 3: Window Class match
    for win in windows:
        cls_lower = win['class'].lower()
        if kw in ["terminal", "wt"] and cls_lower in ["cascadia_hosting_window_class", "consolewindowclass"]:
            return win['hwnd']
        if kw in ["excel"] and "xlmain" in cls_lower:
            return win['hwnd']
        if kw in ["notepad"] and "notepad" in cls_lower:
            return win['hwnd']
        if kw in ["explorer", "file explorer"] and "cabinetwclass" in cls_lower:
            return win['hwnd']

    # Pass 4: Window Title match
    for win in windows:
        title_lower = win['title'].lower()
        if kw in title_lower:
            return win['hwnd']

    return None

def bring_window_to_front(hwnd: int) -> bool:
    """
    Reliably restores and forces the target window to the foreground,
    bypassing Windows foreground-lock restrictions via AttachThreadInput.
    """
    if not hwnd or not win32gui.IsWindow(hwnd):
        return False

    ensure_interactive_desktop()
    try:
        # Check if minimized and restore
        if win32gui.IsIconic(hwnd):
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        else:
            win32gui.ShowWindow(hwnd, win32con.SW_SHOW)

        # Force foreground focus bypassing Windows restriction
        fore_hwnd = win32gui.GetForegroundWindow()
        if fore_hwnd != hwnd:
            fore_thread, _ = win32process.GetWindowThreadProcessId(fore_hwnd)
            cur_thread = win32api.GetCurrentThreadId()
            if fore_thread != cur_thread and fore_thread > 0:
                win32process.AttachThreadInput(cur_thread, fore_thread, True)
                win32gui.SetForegroundWindow(hwnd)
                win32gui.BringWindowToTop(hwnd)
                win32process.AttachThreadInput(cur_thread, fore_thread, False)
            else:
                win32gui.SetForegroundWindow(hwnd)
                win32gui.BringWindowToTop(hwnd)
        time.sleep(0.15)
        return True
    except Exception as e:
        print(f"[Win32] bring_window_to_front error: {e}")
        try:
            win32gui.SetForegroundWindow(hwnd)
            return True
        except Exception:
            return False

APP_REGISTRY = {
    "spotify": {
        "aliases": ["spotify", "music", "songs"],
        "window_keywords": ["spotify"],
        "exec_commands": ["start spotify:", "start spotify"],
        "web_url": "https://open.spotify.com",
        "tab_keywords": ["spotify"]
    },
    "whatsapp": {
        "aliases": ["whatsapp", "wa"],
        "window_keywords": ["whatsapp"],
        "exec_commands": ["start whatsapp:", "start whatsapp"],
        "web_url": "https://web.whatsapp.com",
        "tab_keywords": ["whatsapp"]
    },
    "discord": {
        "aliases": ["discord"],
        "window_keywords": ["discord"],
        "exec_commands": ["start discord:", "start discord"],
        "web_url": "https://discord.com/app",
        "tab_keywords": ["discord"]
    },
    "slack": {
        "aliases": ["slack"],
        "window_keywords": ["slack"],
        "exec_commands": ["start slack:", "start slack"],
        "web_url": "https://app.slack.com",
        "tab_keywords": ["slack"]
    },
    "notion": {
        "aliases": ["notion"],
        "window_keywords": ["notion"],
        "exec_commands": ["start notion:", "start notion"],
        "web_url": "https://www.notion.so",
        "tab_keywords": ["notion"]
    },
    "figma": {
        "aliases": ["figma"],
        "window_keywords": ["figma"],
        "exec_commands": ["start figma:", "start figma"],
        "web_url": "https://www.figma.com",
        "tab_keywords": ["figma"]
    },
    "youtube": {
        "aliases": ["youtube", "yt"],
        "window_keywords": ["youtube"],
        "exec_commands": ["start chrome https://www.youtube.com", "start msedge https://www.youtube.com"],
        "web_url": "https://www.youtube.com",
        "tab_keywords": ["youtube"]
    },
    "github": {
        "aliases": ["github", "git"],
        "window_keywords": ["github"],
        "exec_commands": ["start https://github.com"],
        "web_url": "https://github.com",
        "tab_keywords": ["github"]
    },
    "chatgpt": {
        "aliases": ["chatgpt", "openai"],
        "window_keywords": ["chatgpt"],
        "exec_commands": ["start https://chatgpt.com"],
        "web_url": "https://chatgpt.com",
        "tab_keywords": ["chatgpt"]
    },
    "telegram": {
        "aliases": ["telegram", "tg"],
        "window_keywords": ["telegram"],
        "exec_commands": ["start telegram:", "start telegram"],
        "web_url": "https://web.telegram.org",
        "tab_keywords": ["telegram"]
    },
    "twitter": {
        "aliases": ["twitter", "x"],
        "window_keywords": ["twitter", " x "],
        "exec_commands": ["start https://x.com"],
        "web_url": "https://x.com",
        "tab_keywords": ["twitter", "x.com"]
    },
    "linkedin": {
        "aliases": ["linkedin"],
        "window_keywords": ["linkedin"],
        "exec_commands": ["start https://www.linkedin.com"],
        "web_url": "https://www.linkedin.com",
        "tab_keywords": ["linkedin"]
    },
    "gmail": {
        "aliases": ["gmail", "email", "mail"],
        "window_keywords": ["gmail", "outlook", "mail"],
        "exec_commands": ["start outlookmail:", "start https://mail.google.com"],
        "web_url": "https://mail.google.com",
        "tab_keywords": ["gmail", "inbox", "mail"]
    },
    "vscode": {
        "aliases": ["vscode", "code", "visual studio code"],
        "window_keywords": ["visual studio code", "code"],
        "exec_commands": ["code", "start code"],
        "web_url": "https://vscode.dev",
        "tab_keywords": ["vscode", "visual studio code"]
    },
    "notepad": {
        "aliases": ["notepad", "notes", "text editor"],
        "window_keywords": ["notepad"],
        "exec_commands": ["start notepad"],
        "web_url": None,
        "tab_keywords": []
    },
    "calculator": {
        "aliases": ["calculator", "calc"],
        "window_keywords": ["calculator", "calc"],
        "exec_commands": ["start calc:", "start calc"],
        "web_url": "https://www.google.com/search?q=calculator",
        "tab_keywords": ["calculator"]
    },
    "excel": {
        "aliases": ["excel", "spreadsheet", "sheets"],
        "window_keywords": ["excel"],
        "exec_commands": ["start excel"],
        "web_url": "https://sheets.google.com",
        "tab_keywords": ["excel", "google sheets", "sheets"]
    },
    "word": {
        "aliases": ["word", "doc", "docs", "document"],
        "window_keywords": ["word", "winword"],
        "exec_commands": ["start winword"],
        "web_url": "https://docs.google.com",
        "tab_keywords": ["google docs", "word"]
    },
    "settings": {
        "aliases": ["settings", "system settings"],
        "window_keywords": ["settings"],
        "exec_commands": ["start ms-settings:"],
        "web_url": None,
        "tab_keywords": []
    },
    "terminal": {
        "aliases": ["terminal", "cmd", "command prompt", "powershell", "console", "wt"],
        "window_keywords": ["terminal", "wt", "powershell", "cmd"],
        "exec_commands": ["start wt", "start powershell"],
        "web_url": None,
        "tab_keywords": []
    },
    "explorer": {
        "aliases": ["explorer", "files", "file explorer", "folder"],
        "window_keywords": ["file explorer", "explorer"],
        "exec_commands": ["start explorer"],
        "web_url": None,
        "tab_keywords": []
    },
    "paint": {
        "aliases": ["paint", "mspaint", "draw"],
        "window_keywords": ["paint"],
        "exec_commands": ["start mspaint", "start ms-paint:"],
        "web_url": "https://jspaint.app",
        "tab_keywords": ["jspaint", "paint"]
    },
    "task manager": {
        "aliases": ["task manager", "taskmgr"],
        "window_keywords": ["task manager"],
        "exec_commands": ["start taskmgr"],
        "web_url": None,
        "tab_keywords": []
    },
    "chrome": {
        "aliases": ["chrome", "google chrome"],
        "window_keywords": ["chrome"],
        "exec_commands": ["start chrome"],
        "web_url": "https://www.google.com",
        "tab_keywords": []
    },
    "edge": {
        "aliases": ["edge", "microsoft edge"],
        "window_keywords": ["edge"],
        "exec_commands": ["start msedge"],
        "web_url": "https://www.bing.com",
        "tab_keywords": []
    },
    "browser": {
        "aliases": ["browser", "web browser", "internet"],
        "window_keywords": ["chrome", "edge", "firefox", "brave"],
        "exec_commands": ["start chrome", "start msedge"],
        "web_url": "https://www.google.com",
        "tab_keywords": []
    }
}

def find_browser_tab(keyword: str) -> bool:
    """
    Scans Chrome, Edge, and Firefox browser tabs via UIAutomation.
    If a tab matching keyword is found, activates the window and selects that tab (<30ms).
    """
    if not keyword:
        return False
    try:
        import uiautomation as auto
        root = auto.GetRootControl()
        kw_lower = keyword.lower().strip()
        
        for win in root.GetChildren():
            cls = win.ClassName or ""
            name = win.Name or ""
            # Check for Chrome, Edge, Brave, Firefox top-level windows
            if cls in ("Chrome_WidgetWin_1", "MozillaWindowClass") or any(b in name.lower() for b in ["chrome", "edge", "firefox", "brave"]):
                # Walk tabs
                for elem in win.GetChildren():
                    # Walk elements in tab strip
                    for item in elem.GetChildren():
                        elem_name = (item.Name or "").lower()
                        if kw_lower in elem_name:
                            rect = item.BoundingRectangle
                            if rect:
                                cx = int((rect.left + rect.right) / 2)
                                cy = int((rect.top + rect.bottom) / 2)
                                hwnd = win.NativeWindowHandle
                                if hwnd:
                                    bring_window_to_front(hwnd)
                                time.sleep(0.08)
                                pyautogui.click(cx, cy)
                                time.sleep(0.08)
                                return True
    except Exception as e:
        print(f"[Win32] find_browser_tab note: {e}")
    return False

def resolve_and_open_app(app_query: str, args: str = "", thought_callback = None) -> dict:
    """
    4-Tier Smart App & Web Resolver:
    1. Check if window is already open -> Bring to front (<10ms)
    2. Check if already open in a browser tab -> Switch directly to tab (<30ms)
    3. Launch native installed desktop app cleanly (<150ms)
    4. If not installed / unavailable -> Open official web app URL in browser
    """
    query_clean = app_query.lower().strip().replace("the ", "").replace("app", "").strip()
    if not query_clean:
        return {"success": False, "method": "none", "details": "Empty app query"}

    # Tier 1: Check existing native window directly
    hwnd = find_window_by_name(query_clean, must_be_visible=True)
    if hwnd:
        if thought_callback:
            thought_callback(f"Found active {query_clean.title()} window. Bringing to front...")
        bring_window_to_front(hwnd)
        return {"success": True, "method": "window", "hwnd": hwnd, "details": f"Focused open window '{query_clean}'"}

    # Find matching registry entry
    matched_entry = None
    matched_key = query_clean
    for key, data in APP_REGISTRY.items():
        aliases = data.get("aliases", [])
        if query_clean == key or query_clean in aliases:
            matched_entry = data
            matched_key = key
            break
        words = query_clean.split()
        if any((a in words or f" {a} " in f" {query_clean} ") for a in aliases if len(a) > 1):
            matched_entry = data
            matched_key = key
            break

    # Tier 2: Check open browser tabs
    tab_search_keywords = (matched_entry.get("tab_keywords", []) + [query_clean]) if matched_entry else [query_clean]
    for tab_kw in tab_search_keywords:
        if tab_kw:
            if thought_callback:
                thought_callback(f"Checking browser tabs for '{tab_kw}'...")
            found_tab = find_browser_tab(tab_kw)
            if found_tab:
                return {"success": True, "method": "tab", "details": f"Switched to browser tab matching '{tab_kw}'"}

    # Tier 3: Try launching native app cleanly (pick single best executable without duplicate loops)
    exec_commands = matched_entry.get("exec_commands", []) if matched_entry else [f"start {query_clean} {args}".strip()]
    cmd = exec_commands[0] if exec_commands else f"start {query_clean}"

    try:
        if thought_callback:
            thought_callback(f"Launching {matched_key.title()} desktop app...")
        full_cmd = f"{cmd} {args}".strip() if args else cmd
        subprocess.Popen(full_cmd, shell=True)

        # Poll up to 2.5s for window to spawn and register
        for _ in range(12):
            time.sleep(0.2)
            hwnd = find_window_by_name(matched_key) or find_window_by_name(query_clean)
            if hwnd:
                bring_window_to_front(hwnd)
                return {"success": True, "method": "native", "hwnd": hwnd, "details": f"Launched native app via '{cmd}'"}
    except Exception as le:
        print(f"[Win32] Native launch note for {cmd}: {le}")

    # Fallback to secondary command if defined (e.g. powershell if wt was missing)
    if len(exec_commands) > 1:
        fallback_cmd = exec_commands[1]
        try:
            full_fallback = f"{fallback_cmd} {args}".strip() if args else fallback_cmd
            subprocess.Popen(full_fallback, shell=True)
            for _ in range(10):
                time.sleep(0.2)
                hwnd = find_window_by_name(matched_key) or find_window_by_name(query_clean)
                if hwnd:
                    bring_window_to_front(hwnd)
                    return {"success": True, "method": "native", "hwnd": hwnd, "details": f"Launched native app via '{fallback_cmd}'"}
        except Exception:
            pass

    # Tier 4: Fallback to Web App URL in browser
    web_url = matched_entry.get("web_url") if matched_entry else None
    if not web_url and query_clean not in ["notepad", "settings", "terminal", "explorer", "task manager", "cmd", "powershell"]:
        web_url = f"https://www.google.com/search?q={query_clean}"

    if web_url:
        if thought_callback:
            thought_callback(f"App not installed locally. Opening official web app: {web_url}")
        subprocess.Popen(f'start "" "{web_url}"', shell=True)
        return {"success": True, "method": "web", "url": web_url, "details": f"Opened web fallback '{web_url}'"}

    return {"success": False, "method": "none", "details": f"Could not locate or launch '{app_query}'"}

def launch_or_focus_app(app_name: str, args: str = "") -> int:
    """
    Finds existing window for app_name and focuses it; if not found, uses resolve_and_open_app.
    Returns the HWND of the app or 0.
    """
    res = resolve_and_open_app(app_name, args)
    return res.get("hwnd", 0)

def send_win32_background_text(hwnd: int, text: str) -> bool:
    """
    Sends text silently to a window's message queue without taking over the physical mouse.
    """
    if not hwnd or not win32gui.IsWindow(hwnd):
        return False

    ensure_interactive_desktop()
    try:
        targets = [hwnd]
        def enum_child(child, _):
            targets.append(child)
            return True
        try:
            win32gui.EnumChildWindows(hwnd, enum_child, None)
        except Exception:
            pass

        for target in targets:
            try:
                win32api.SendMessage(target, win32con.WM_ACTIVATE, win32con.WA_ACTIVE, 0)
                win32api.SendMessage(target, win32con.WM_SETFOCUS, 0, 0)
                for char in text:
                    if char == '\n':
                        win32api.PostMessage(target, win32con.WM_KEYDOWN, win32con.VK_RETURN, 1)
                        win32api.PostMessage(target, win32con.WM_KEYUP, win32con.VK_RETURN, 0xC0000001)
                    elif char == '\t':
                        win32api.PostMessage(target, win32con.WM_KEYDOWN, win32con.VK_TAB, 1)
                        win32api.PostMessage(target, win32con.WM_KEYUP, win32con.VK_TAB, 0xC0000001)
                    else:
                        vk_code = win32api.VkKeyScan(char) & 0xFF
                        win32api.PostMessage(target, win32con.WM_KEYDOWN, vk_code, 1)
                        win32api.PostMessage(target, win32con.WM_CHAR, ord(char), 1)
                        win32api.PostMessage(target, win32con.WM_KEYUP, vk_code, 0xC0000001)
                    time.sleep(0.002)
            except Exception:
                pass
        return True
    except Exception as e:
        print(f"[Win32] Background text dispatch failed: {e}")
        return False

def get_active_screen_pdf_files() -> tuple[list[str], str]:
    """
    Dynamically discovers PDF files from the folder currently open on the user's screen.
    Prioritizes:
    1. Active Windows File Explorer folder / tabs via Shell COM API
    2. Selected PDF items in the open File Explorer window
    3. Window titles matching local directories
    4. Fallback to demo_invoices or workspace folders
    """
    import os
    import glob

    # 1. Shell COM API - Inspect open Explorer folder paths & selections
    try:
        import pythoncom
        import win32com.client
        import win32gui

        pythoncom.CoInitialize()
        shell = win32com.client.Dispatch("Shell.Application")
        for w in shell.Windows():
            try:
                folder_path = w.Document.Folder.Self.Path
                if folder_path and os.path.isdir(folder_path):
                    # Check if user highlighted/selected specific PDF files
                    selected_pdfs = []
                    try:
                        for item in w.Document.SelectedItems():
                            if item.Path.lower().endswith(".pdf"):
                                selected_pdfs.append(item.Path)
                    except Exception:
                        pass
                    if selected_pdfs:
                        return selected_pdfs, f"selected PDFs in '{os.path.basename(folder_path)}'"

                    # All PDFs in this specific open folder
                    folder_pdfs = [
                        os.path.join(folder_path, f)
                        for f in os.listdir(folder_path)
                        if f.lower().endswith(".pdf")
                    ]
                    if folder_pdfs:
                        return folder_pdfs, f"open folder '{os.path.basename(folder_path)}'"
            except Exception:
                continue
    except Exception as e:
        print(f"[Win32] Shell folder inspection note: {e}")

    # 2. Window Title matching for active directories
    try:
        import win32gui
        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        open_titles = []
        def enum_win(hwnd, _):
            if win32gui.IsWindowVisible(hwnd):
                t = win32gui.GetWindowText(hwnd).strip()
                if t:
                    open_titles.append(t.lower())
        win32gui.EnumWindows(enum_win, None)

        for title in open_titles:
            candidate = os.path.join(root_dir, title)
            if os.path.isdir(candidate):
                pdfs = glob.glob(os.path.join(candidate, "*.pdf"))
                if pdfs:
                    return pdfs, f"open folder '{title}'"
            candidate_home = os.path.join(os.path.expanduser("~"), title)
            if os.path.isdir(candidate_home):
                pdfs = glob.glob(os.path.join(candidate_home, "*.pdf"))
                if pdfs:
                    return pdfs, f"open folder '{title}'"
    except Exception:
        pass

    # 3. Fallback to workspace invoice folders
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for candidate_dir in [
        os.path.join(root_dir, "demo_invoices"),
        os.path.join(root_dir, "invoices"),
        os.path.join(os.getcwd(), "demo_invoices"),
        os.getcwd()
    ]:
        if os.path.isdir(candidate_dir):
            pdfs = glob.glob(os.path.join(candidate_dir, "*.pdf"))
            if pdfs:
                return pdfs, f"workspace folder '{os.path.basename(candidate_dir)}'"

    return [], "no PDF files found"

def inject_clipboard_text(text: str, press_enter: bool = False, target_hwnd: int = 0) -> bool:
    """
    Copies text to the system clipboard, focuses target_hwnd if provided, and sends Ctrl+V.
    """
    try:
        if target_hwnd:
            bring_window_to_front(target_hwnd)
            time.sleep(0.1)

        pyperclip.copy(text)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.1)
        if press_enter:
            pyautogui.press('enter')
        return True
    except Exception as e:
        print(f"[Win32] inject_clipboard_text failed: {e}")
        return False

def inject_excel_table(rows: list, headers: list = None, target_sheet_name: str = None, output_filename: str = "extracted_invoices.xlsx") -> bool:
    """
    Guaranteed visible Excel table injection.
    Creates/updates the .xlsx file on disk, formats headers and columns,
    and opens it in Microsoft Excel in the foreground so the user visually sees the result!
    """
    ensure_interactive_desktop()
    file_path = os.path.abspath(output_filename)
    if os.path.exists(file_path):
        try:
            with open(file_path, "a+"):
                pass
        except IOError:
            ts = time.strftime("%H%M%S")
            file_path = os.path.abspath(f"extracted_invoices_{ts}.xlsx")

    # 1. Generate formatted Excel workbook on disk via openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = target_sheet_name or "Invoices"

        if headers:
            ws.append(headers)
            # Format Header Row: Bold, dark fill, centered
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Append data rows
        for r in rows:
            if isinstance(r, dict):
                col_values = list(r.values())
            elif isinstance(r, (list, tuple)):
                col_values = list(r)
            else:
                col_values = [str(r)]
            ws.append(col_values)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(file_path)
        print(f"[Win32] Saved formatted Excel table to: {file_path}")

        # Physically launch Excel with the saved file open on screen!
        os.startfile(file_path)
        time.sleep(0.5)
        return True
    except Exception as e:
        print(f"[Win32] openpyxl file launch error ({e}), falling back to COM API...")

    # 2. COM API Fallback
    try:
        import win32com.client
        excel = None
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            excel = win32com.client.Dispatch("Excel.Application")

        if excel:
            excel.Visible = True
            wb = excel.ActiveWorkbook or excel.Workbooks.Add()
            sheet = wb.ActiveSheet
            start_row = 1
            if headers:
                for col_idx, h in enumerate(headers, 1):
                    sheet.Cells(1, col_idx).Value = str(h)
                    sheet.Cells(1, col_idx).Font.Bold = True
                start_row = 2
            for row_idx, r in enumerate(rows, start_row):
                vals = list(r.values()) if isinstance(r, dict) else list(r) if isinstance(r, (list, tuple)) else [str(r)]
                for col_idx, val in enumerate(vals, 1):
                    sheet.Cells(row_idx, col_idx).Value = str(val)
            sheet.Columns.AutoFit()
            bring_window_to_front(excel.Hwnd)
            return True
    except Exception as ce:
        print(f"[Win32] Excel COM fallback error: {ce}")

    # 3. TSV Clipboard fallback
    tsv_lines = []
    if headers:
        tsv_lines.append("\t".join([str(h) for h in headers]))
    for r in rows:
        vals = list(r.values()) if isinstance(r, dict) else list(r) if isinstance(r, (list, tuple)) else [str(r)]
        tsv_lines.append("\t".join([str(v) for v in vals]))
    formatted_tsv = "\n".join(tsv_lines) + "\n"

    excel_hwnd = find_window_by_name("excel")
    return inject_clipboard_text(formatted_tsv, press_enter=False, target_hwnd=excel_hwnd)
